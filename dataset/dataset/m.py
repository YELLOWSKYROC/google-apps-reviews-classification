import openpyxl

workbookPath = "/Users/htp/Desktop/Year3_Project.xlsx"
sheetName = 'Sheet1'
workbook = openpyxl.load_workbook(workbookPath)
sheet = workbook.get_sheet_by_name(sheetName)

for i in range(2,2002):
    if sheet.cell(i, 3).value == 1:
        sheet.cell(i, 3).value = "bug report"
    if sheet.cell(i, 3).value == 2:
        sheet.cell(i, 3).value = "feature request"
    if sheet.cell(i, 3).value == 3:
        sheet.cell(i, 3).value = "User experience"
    if sheet.cell(i, 3).value == 4:
        sheet.cell(i, 3).value = "rating"

workbook.save(workbookPath)
workbook.close()
